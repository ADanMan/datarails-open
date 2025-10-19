from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from app import excel_loader


@pytest.fixture()
def workbook_path(tmp_path: Path) -> Path:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Actuals"
    ws1.append(["period", "department", "account", "value", "currency", "metadata"])
    ws1.append(["2024-01", "Sales", "Revenue", 1000, "USD", "Q1 actuals"])

    ws2 = wb.create_sheet("Budget")
    ws2.append(["period", "department", "account", "value", "currency"])
    ws2.append(["2024-01", "Sales", "Revenue", 1200, "USD"])
    ws2.append(["2024-01", "Marketing", "Spend", -300, "USD"])

    path = tmp_path / "financials.xlsx"
    wb.save(path)
    wb.close()
    return path


def test_read_workbook_multiple_sheets(workbook_path: Path) -> None:
    rows = excel_loader.read_workbook(workbook_path, sheets=["Actuals", "Budget"])
    assert len(rows) == 3
    assert rows[0] == ("2024-01", "Sales", "Revenue", 1000.0, "USD", "Q1 actuals")
    assert ("Marketing", "Spend") in {(row[1], row[2]) for row in rows}


def test_read_workbook_missing_required_columns(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["period", "department", "account"])  # missing value column
    ws.append(["2024-01", "Sales", "Revenue"])

    path = tmp_path / "missing.xlsx"
    wb.save(path)
    wb.close()

    with pytest.raises(ValueError, match=r"Missing required columns: \['value'\]"):
        excel_loader.read_workbook(path)


def test_load_workbook_file_inserts_rows(tmp_path: Path, sqlite_connection) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["period", "department", "account", "value", "currency"])
    ws.append(["2024-02", "Finance", "Cost", -150])

    table = Table(displayName="FinanceTable", ref="A1:E2")
    ws.add_table(table)

    path = tmp_path / "table.xlsx"
    wb.save(path)
    wb.close()

    summary = excel_loader.load_workbook_file(
        sqlite_connection,
        path,
        source="excel-upload",
        scenario="plan",
        tables=["FinanceTable"],
    )
    assert summary.rows_loaded == 1

    cursor = sqlite_connection.execute(
        "SELECT period, department, account, value, currency FROM financial_facts"
    )
    stored = [tuple(row) for row in cursor.fetchall()]
    assert stored == [("2024-02", "Finance", "Cost", -150.0, "USD")]
