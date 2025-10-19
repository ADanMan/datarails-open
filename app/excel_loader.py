"""Excel ingestion helpers for the datarails-open MVP."""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Sequence

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from . import database
from .loader import LoadSummary, REQUIRED_COLUMNS, _normalise_row


def _normalised_rows_from_iterable(
    header_row: Sequence[object],
    data_rows: Iterable[Sequence[object]],
) -> List[tuple[str, str, str, float, str, str]]:
    headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
    if not any(headers):
        return []
    missing = set(REQUIRED_COLUMNS) - {header for header in headers if header}
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    normalised: List[tuple[str, str, str, float, str, str]] = []
    for row in data_rows:
        if not any(cell not in (None, "") for cell in row):
            continue
        as_dict = {
            header: "" if value is None else str(value)
            for header, value in zip(headers, row)
            if header
        }
        normalised.append(_normalise_row(as_dict))
    return normalised


def _rows_from_sheet(sheet: Worksheet) -> List[tuple[str, str, str, float, str, str]]:
    rows_iter = sheet.iter_rows(values_only=True)
    for header_row in rows_iter:
        if header_row and any(cell not in (None, "") for cell in header_row):
            break
    else:
        return []
    return _normalised_rows_from_iterable(header_row, rows_iter)


def _rows_from_table(sheet: Worksheet, table_name: str) -> List[tuple[str, str, str, float, str, str]]:
    table = sheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows_iter = sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )
    header_row = next(rows_iter, None)
    if header_row is None:
        return []
    return _normalised_rows_from_iterable(header_row, rows_iter)


def read_workbook(
    path: Path | str,
    *,
    sheets: Sequence[str] | None = None,
    tables: Sequence[str] | None = None,
) -> List[tuple[str, str, str, float, str, str]]:
    """Read an Excel workbook and return normalised tuples ready for insertion."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() not in {".xlsx"}:
        raise ValueError("Only .xlsx files are supported by the Excel loader")

    workbook = load_workbook(path, data_only=True)
    try:
        requested_sheets = list(sheets) if sheets else workbook.sheetnames
        for sheet_name in requested_sheets:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{sheet_name}' not found in workbook")

        table_index = {
            name: ws
            for ws in workbook.worksheets
            for name in ws.tables
        }
        if tables:
            for table_name in tables:
                if table_name not in table_index:
                    raise ValueError(f"Table '{table_name}' not found in workbook")

        normalised_rows: List[tuple[str, str, str, float, str, str]] = []
        if tables:
            for table_name in tables:
                ws = table_index[table_name]
                if sheets and ws.title not in sheets:
                    raise ValueError(
                        f"Table '{table_name}' is located on worksheet '{ws.title}', "
                        "which is not in the selected sheets",
                    )
                normalised_rows.extend(_rows_from_table(ws, table_name))
        else:
            for sheet_name in requested_sheets:
                sheet = workbook[sheet_name]
                normalised_rows.extend(_rows_from_sheet(sheet))
        return normalised_rows
    finally:
        workbook.close()


def load_workbook_file(
    conn,
    path: Path | str,
    *,
    source: str,
    scenario: str,
    sheets: Sequence[str] | None = None,
    tables: Sequence[str] | None = None,
) -> LoadSummary:
    rows = read_workbook(path, sheets=sheets, tables=tables)
    payload = [
        (source, scenario, period, department, account, value, currency, metadata)
        for period, department, account, value, currency, metadata in rows
    ]
    inserted = database.insert_rows(conn, payload)
    return LoadSummary(rows_loaded=inserted, source=source, scenario=scenario)
